"""
server.py  —  Rapid Dashboard Backend (Python + FastAPI)

PERFORMANCE FIX:
  All Excel sheets are read ONCE at startup and cached as plain
  Python dicts in memory. Every API request is then just a
  dictionary lookup — no file I/O, no XML parsing per request.
  This matches the speed of the original JavaScript backend.

Install:
    pip install fastapi uvicorn[standard] openpyxl

Run:
    python server.py
"""

import openpyxl
import uvicorn
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware

# ════════════════════════════════════════════════════════════
#  STARTUP — read entire Excel into memory once
# ════════════════════════════════════════════════════════════
EXCEL_PATH = "data.xlsx"

# This dict holds ALL sheet data as plain Python lists of dicts
# Structure: { "sheet_name": [ {col: val, ...}, ... ] }
SHEET_CACHE: dict[str, list[dict]] = {}

def load_excel_to_cache():
    print(f"📂 Loading {EXCEL_PATH} into memory...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except FileNotFoundError:
        print(f"❌ ERROR: {EXCEL_PATH} not found. Place it next to server.py")
        return

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows  = list(sheet.iter_rows(values_only=True))

        if not rows:
            SHEET_CACHE[sheet_name] = []
            continue

        # First row = headers
        headers = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(rows[0])
        ]

        records = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue  # skip blank rows
            records.append({
                headers[i]: (str(v) if v is not None else "")
                for i, v in enumerate(row)
            })

        SHEET_CACHE[sheet_name] = records
        print(f"   ✅ '{sheet_name}' — {len(records):,} rows cached")

    wb.close()
    print(f"🚀 All sheets loaded. Serving from memory.\n")

# Load immediately at module import (before first request)
load_excel_to_cache()


# ════════════════════════════════════════════════════════════
#  FASTAPI APP
# ════════════════════════════════════════════════════════════
app = FastAPI(title="Rapid Dashboard API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── GET /api/node/{node_id} ──────────────────────────────────
# Returns all rows for a sheet — served from cache, instant
@app.get("/api/node/{node_id}")
def get_node_data(node_id: str):
    return SHEET_CACHE.get(node_id, [])


# ── GET /api/summary/{node_id} ───────────────────────────────
# Returns total/success/failed counts — served from cache, instant
@app.get("/api/summary/{node_id}")
def get_summary(node_id: str):
    data = SHEET_CACHE.get(node_id, [])

    total    = len(data)
    success  = sum(1 for d in data if d.get("Status", "").upper() == "SUCCESS")
    failed   = sum(1 for d in data if d.get("Status", "").upper() == "FAILED")
    critical = sum(1 for d in data if d.get("CRITICAL", "").upper() == "YES")

    return {
        "total":    total,
        "success":  success,
        "failed":   failed,
        "critical": critical,
    }


# ── GET /api/health ──────────────────────────────────────────
@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "excel":  EXCEL_PATH,
        "sheets": [
            {"name": name, "rows": len(rows)}
            for name, rows in SHEET_CACHE.items()
        ],
    }


# ── Run ──────────────────────────────────────────────────────
if __name__ == "__main__":
    uvicorn.run("server:app", host="0.0.0.0", port=5000, reload=False)
    #                                                      ^^^^^^^^^^^
    # reload=False is important for performance —
    # reload=True watches files and re-imports the module,
    # which means re-reading the Excel on every code change.
    # Turn it True only when actively editing server.py.
